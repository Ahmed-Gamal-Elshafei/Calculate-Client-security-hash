import openpyxl


def get_item(bot_id, file_path="../data3.xlsx"):
    """
    Get the first item matching the criteria from the Excel sheet.
    """

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Map headers to indices
        headers = {sheet.cell(row=1, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1)}

        # Iterate over rows to find a matching item
        for row_idx in range(2, sheet.max_row + 1):
            _status = sheet.cell(row=row_idx, column=headers["_status"]).value
            retry_number = sheet.cell(row=row_idx, column=headers["retry_number"]).value
            lock = sheet.cell(row=row_idx, column=headers["lock"]).value
            _type = sheet.cell(row=row_idx, column=headers["Type"]).value

            if (_status in ["wait", "failed"]) and retry_number < 3 and not lock and _type == "WI5":
                # wait random time before pick for 0 to 500 milliseconds
                import random
                import time
                time.sleep(random.randint(0, 500) / 1000)
                # Lock the item
                sheet.cell(row=row_idx, column=headers["lock"]).value = bot_id
                workbook.save(file_path)

                # Return the item as a dictionary
                item = {header: sheet.cell(row=row_idx, column=col_idx).value for header, col_idx in headers.items()}
                workbook.close()
                return item, row_idx

        workbook.close()
    except Exception as e:
        print(f"Error in get_item: {e}")
    return None, None


def update_item(row_idx, updated_values, file_path="../data3.xlsx"):
    """
    Updates a specific row in the Excel file.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Map headers to indices
        headers = {sheet.cell(row=1, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1)}

        # Update the row with new values
        for column_name, new_value in updated_values.items():
            if column_name in headers:
                col_idx = headers[column_name]
                sheet.cell(row=row_idx, column=col_idx).value = new_value

        workbook.save(file_path)
        print("Item updated successfully.")
    except Exception as e:
        print(f"Error in update_item: {e}")


def check_item_bot(bot_id, row_idx):
    """
    Checks if an item is locked by a specific bot.
    """
    file_path = "../data3.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Map headers to indices
        headers = {sheet.cell(row=1, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1)}

        # Check if the item is locked by the specified bot
        lock_value = sheet.cell(row=row_idx, column=headers["lock"]).value
        return lock_value == bot_id

    except Exception as e:
        print(f"Error in check_item_bot: {e}")
    return False
